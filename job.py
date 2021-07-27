import re
import time
import requests
import openpyxl
import tldextract
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.utils import ChromeType


job_domain = "javascript"
location = "bangalore"
url = 'https://www.naukri.com/%s-jobs-in-%s?k=%s&l=%s' % (job_domain , location , job_domain , location)
max_page = 3

wb = Workbook()
sheet =  wb.active
sheet.title = "JOBS"
col = 11
l = ['Company' , 'JobTitle' , 'Experience' ,'Salary' , 'Location' , 'Skills Required' , 'Job Link' , 'Job Description' , 'Role' , 'Industry' , 'Employment Type']
for i in range(1 , col+1):
    c = sheet.cell(row = 1, column = i)
    c.value = l[i-1]
wb.save("/home/pathreaig/naukri-scraper/demo.xlsx")

driver = webdriver.Chrome(ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install())
driver.get(url)

job_title = []
job_link = []
comp_title = []
job_desc = []
sals = []
skls = []
tipu = []
od_details = []
i=0

for nex in range(2,max_page):
    time.sleep(3)

    src = driver.page_source
    soup = BeautifulSoup(src , 'html.parser')
    results = soup.find('section' , class_ = 'listContainer fleft')

    time.sleep(3)
    comp_name = results.find_all('a' , class_ = 'subTitle ellipsis fleft')
    j_title = results.find_all('a' , class_ = 'title fw500 ellipsis')
    salaries = results.find_all('span' , attrs={'class': re.compile('^ellipsis fleft fs12 lh16.*')})
    #print("j_title: ",j_title)



    for i in range(len(comp_name)):
        comp_title.append(comp_name[i].text)
        job_title.append(j_title[i].text)
        
    
    for i in range(len(salaries)):
        sals.append(salaries[i].text)
        
    for i in range(len(j_title)):
        job_link.append(j_title[i]['href'])

    #print("sals now is: ",sals)

    time.sleep(3)
    skills = results.find_all('ul' , class_ = 'tags has-description')

    for index  , skill in enumerate(skills):
        nn = skill.find_all('li' , class_ = 'fleft fs12 grey-text lh16 dot')
        for jj in nn:
            skls.append(jj.text)
        tipu.append(str(skls))
        skls = []

    wb.save("/home/pathreaig/naukri-scraper/demo.xlsx")
    driver.find_element_by_link_text(str(nex)).click()
    time.sleep(0.5)

for link in job_link:
    ext = tldextract.extract(link)
    jd1 = []
    jd2 = []
    
    time.sleep(0.5)
    driver.get(link)
    src=driver.page_source
    soup = BeautifulSoup(src,'html.parser')
    results = soup.find('section', attrs={'class': re.compile('\W*(job-desc)\W*')})
    #print('link is: ',link)
    #od_details.append(link)
    if(results == None):
         
         jd1 = soup.find('div', class_ = 'clearboth description')
         #print('jd1 result len is:', len(jd1)) 
         job_desc.append(jd1.text)
         #od1 = jd1.find_all('p', class_ = 'coPE getRoleLabel')
         for i in range(0,13):
             od_details.append(i)
             #od1_label = od1[i].find('em')
             #od1_span = od1[i].find('span')
             #od_details.append(od1_label,od1_span)
    
    else:
         jd2 = results.find_all('div', class_ = 'dang-inner-html')
         #print('len jd2 is: ', len(jd2))
         for i in range(len(jd2)):
             #print(jd2[i].text,'\n')
             job_desc.append(jd2[i].text)
         od2 = results.find_all('div' , class_ = 'details')
         for i in range(len(od2)):
             #print('od2 : ', i, ' is ', od2[i])
             od2_label = od2[i].find_all('label')
             od2_a = od2[i].find_all('a')
             od2_span = od2[i].find_all('span')

             #o_details.append(od2_item)
             if(len(od2_a)>0):
                 for i in range(len(od2_label)):
                     #print(od2_label[i].text)
                     #print(od2_a[i].text)
                     od_details.append(od2_label[i].text)
                     od_details.append(od2_a[i].text)
             elif(len(od2_span)>0):
                 for i in range(len(od2_span)):
                     #print('span i ', i , ' ', od2_span[i])
                     od_details.append(od2_span[i].text)
             
            



#print('od_details now is: ',od_details)
    #print('jd2:',jd2)	
    
    
            
    
        
    #job_desc.append(jd.text)

time.sleep(0.3)


row = len(comp_title)
col = 5
#print(row)
#print(len(job_title) , len(comp_title) , len(sals) , len(tipu))

sheet =  wb.active

k = 0
j = 0
for i in range(2 , row+2):
    c1 = sheet.cell(row = i , column = 1) #company
    c1.value = comp_title[i-2]

    c1 = sheet.cell(row = i , column = 2) #job_title
    c1.value = job_title[i-2]

    c1 = sheet.cell(row = i , column = 3) #experience
    #print("k is: ",k,'length is: ',len(sals))
    if k<len(sals):
        c1.value = sals[k]
    
    c1 = sheet.cell(row = i , column = 4) #salary
    if k<len(sals)-1:
        c1.value = sals[k+1]
    
    c1 = sheet.cell(row = i , column = 5) #location
    if k<len(sals)-2:
        c1.value = sals[k+2]
    
    k+=3
    #print('j now is: ',j, 'length of od_details: ', len(od_details))
    c1 = sheet.cell(row = i , column = 6) #skills
    c1.value = tipu[i-2]

    c1 = sheet.cell(row = i , column = 7) #job_link
    c1.value = job_link[i-2]

    c1 = sheet.cell(row = i, column = 8) #job description
    c1.value = job_desc[i-2]

    c1 = sheet.cell(row = i , column = 9) #role
    if(j<len(od_details)):
        #print('j is: ',j)
        c1.value = od_details[j+1]

    c1 = sheet.cell(row = i , column = 10) #Industry
    if(j<len(od_details)):
        #print('j is: ',j)
        c1.value = od_details[j+3]
    

    c1 = sheet.cell(row = i , column = 11) #Employment Type
    if(j<len(od_details)):
        #print('j is: ',j)
        c1.value = od_details[j+6]
    

    j+=13
    
wb.save("/home/pathreaig/naukri-scraper/demo.xlsx")
